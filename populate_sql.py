import os
import django
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

from movies.models import Movies


def populate():
    movie_data = read_xlsx()
    add_movie(movie_data)


def add_movie(data):
    movie_list = [
        Movies(id=item['id'], title=item['title'], director=item['director'], date=item['date'], actor=item['actor'],
               score=item['score'], intro=item['intro'], tags=item['tags'], imdb=item['imdb'],
               time_length=item['time_length'], pic=item['pic']) for item in data]
    m = Movies.objects.get_or_create()[0]
    m.save()
    Movies.objects.bulk_create(movie_list)
    return movie_list


def read_xlsx():
    # 加载xlsx文件
    workbook = load_workbook(filename='movies.xlsx')

    # 选择第一个工作表
    sheet = workbook.worksheets[0]

    # 获取表头
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)

    # 遍历每一行，将数据存入字典
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {}
        for i, value in enumerate(row):
            item[headers[i]] = value
        data.append(item)
    return data


if __name__ == '__main__':
    print('Starting Rango population script...')
    populate()
